from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook("sales.xlsx")
sheet = wb.active
sheet["E1"] = "Revenue"

print("Sheet title:", sheet.title)

rows = list(sheet.iter_rows(values_only = True))

header = rows[0]
data_rows = rows[1:]

for cell in sheet[1]:
    cell.font = Font(bold=True)

print("Header", header)

for row_idx in range(2, sheet.max_row + 1):
    qty = sheet[f"C{row_idx}"].value
    price = sheet[f"D{row_idx}"].value
    sheet[f"E{row_idx}"] = qty * price

for column in sheet.columns :
    max_lenght = 0
    col_letter = column[0].column_letter

    for cell in column:
        if cell.value:
            max_lenght = max(max_lenght, len(str(cell.value)))

    sheet.column_dimensions[col_letter].width = max_lenght + 2

summary_sheet = wb.create_sheet(title="Summary")
summary_sheet["A1"] = "Product"
summary_sheet["B1"] = "Total Revenue"

for cell in summary_sheet[1]:
    cell.font = Font(bold=True)

revenue_by_product = {}
for row_idx in range(2, sheet.max_row + 1):
    product = sheet[f"B{row_idx}"].value
    revenue = sheet[f"E{row_idx}"].value

    revenue_by_product[product] = revenue_by_product.get(product, 0)+revenue

row_num = 2
for product, total_rev in revenue_by_product.items():
    summary_sheet[f"A{row_num}"] = product
    summary_sheet[f"B{row_num}"] = round(total_rev, 2)
    row_num += 1







wb.save("sales_with_revenue.xlsx")
print("Saved: sales_with revenue.xlsx")
    
